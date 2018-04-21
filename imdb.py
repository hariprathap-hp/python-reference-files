# -*- coding: utf-8 -*-
"""
Created on Wed Apr 11 14:19:29 2018
@author: Hari Prathap
"""
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import scrapy, json, re, os, openpyxl
js_file = " "
oscar_2018 = {"Oscars Categories":"Oscar Winners"}
oscar_dict = {}
urls = []

class ParseIMDB(scrapy.Spider):
    name = 'imdb'
    curr_year = " "
    
    def start_requests(self):   
        '''for year in  range(1929,2018):
            curr_year = year
            dyn_url = "https://www.imdb.com/event/ev0000003/" + str(year) + "/1/?ref_=ev_eh"
            print(dyn_url)
            urls.append(dyn_url)'''
            
        urls = ['https://www.imdb.com/event/ev0000003/2018/1/?ref_=ev_eh']
                 
        for url in urls:
            yield scrapy.Request(url=url, callback=self.parse_imdb)
            
    def parse_imdb(self,response):
        my_file = open("C:\\Users\\Aricent\\tutorial\\javascript.json",'wb+')
        js_output = response.xpath('//span[@class="ab_widget"]//script[@type="text/javascript"]').extract()
        oscar_year = response.xpath('//div[@class="event-year-header__year"]//text()').extract_first()
        print(oscar_year)
        my_file.write(js_output[0].encode("utf-8"))
        my_file.close()
        self.extract_movie_info(oscar_year)
            
    def extract_movie_info(self, oscar_year):
        print("In function")
        json_file = open("C:\\Users\\Aricent\\tutorial\\javascript.json")
        my_json_out = open("C:\\Users\\Aricent\\tutorial\\final_json.json","w+")
        my_lines = json_file.readlines()
        
        #print(my_lines)
        my_rex = re.compile(r'(?<=\'center-(\d)-react\',)(.*)(?=]\);)') 
        for line in my_lines:
            if "nomineesWidgetModel" in line:      
                my_line = my_rex.search(line)
                my_json_out.write(my_line.group())
        my_file_1 = open("C:\\Users\\Aricent\\tutorial\\final_json.json")
        json_data = json.load(my_file_1)
        length = len(json_data["nomineesWidgetModel"]["eventEditionSummary"]["awards"][0]["categories"])
        print(length)
        
        for i in range(0,length):
            if json_data["nomineesWidgetModel"]["eventEditionSummary"]["awards"][0]["categories"][i]["nominations"][0]["isWinner"]:
                try:
                    categories = (json_data["nomineesWidgetModel"]["eventEditionSummary"]["awards"][0]["categories"][i]["categoryName"]).upper()
                    winners = json_data["nomineesWidgetModel"]["eventEditionSummary"]["awards"][0]["categories"][i]["nominations"][0]["primaryNominees"][0]["name"]
                except IndexError:
                    print("Winner's Name is not published")
                oscar_2018[categories] = winners
        self.populate_excel(oscar_2018,oscar_year)
        
        #print(oscar_2018)
        my_file_1.close()
        json_file.close()
        my_json_out.close()
        
        #closing the files        
        os.remove("C:\\Users\\Aricent\\tutorial\\javascript.json")
        os.remove("C:\\Users\\Aricent\\tutorial\\final_json.json")        
        
    def populate_excel(self, oscar_dict, oscar_year):
        YellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        excel = openpyxl.Workbook() #Open the empty excel workbook
        active_sheet = excel.create_sheet(oscar_year)
        
        active_sheet['A1'].fill = YellowFill
        active_sheet['B1'].fill = YellowFill 
        
        category = 1
        winner = 1
                
        for key, value in oscar_dict.items():
            ex_category = "A" + str(category)
            ex_winner = "B" + str(winner)            
            active_sheet[ex_category] = key
            active_sheet[ex_winner] = value
            category = category+1
            winner = winner+1
        
        excel.save("C:\\Users\\Aricent\\tutorial\\tutorial\\spiders\\Oscars_winners.xlsx")